from ris import pysqldb
from openpyxl import load_workbook
import datetime
import glob
import os
import configparser
import pandas as pd
import webbrowser

timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
ts = datetime.datetime.now().strftime('%Y-%m-%d')
year = datetime.datetime.now().strftime('%Y')

LOC = r'E:\RIS\Staff Folders\Seth H\Misc\srts_geotab'

#pathlib.Path(glob.glob(os.path.join(r'%s' % ("\\\\" + data['PATHS']['%s' % i].lstrip('\\')),"*.xlsx"))[0]).stem

def read_excel(path,df=False):
    """
    Function that reads in excel files.
    :param path (str): Path of excel file to read in.
    :param df (boolean):
    :return: tuple containg list of dataframes and sheetnames.
    """

    wb = load_workbook(path,read_only=True)
    tables = []
    sheets = wb.get_sheet_names()

    if df:
       for i in sheets:
        data = pd.DataFrame(wb['%s' % i].values)
        new_header = data.iloc[0]
        tbl = data[1:]
        tbl.columns = new_header
        tables.append(tbl)

    return tables, sheets


files = glob.glob(os.path.join(LOC,"*_results.xlsx")) #searching folder path for all files containing _result.xlsx ending
master = []

for i in files:
    master.append(read_excel(i, df=True)[0][0])

master_file = pd.concat(master, ignore_index=False) #combining all files into masterfile

master_file.to_excel('E:\RIS\Staff Folders\Samuel\Requests\SH\MISC\master_file.xlsx',index=False)